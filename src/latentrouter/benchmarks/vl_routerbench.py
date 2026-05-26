from __future__ import annotations

import re
import json
import pickle
import tarfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from latentrouter.benchmarks.base import BenchmarkAdapter, BenchmarkProtocol
from latentrouter.benchmarks.common import (
    lambda_grid,
    materialize_routes,
    oracle_frontier,
    single_model_frontier,
    write_evaluation_outputs,
)
from latentrouter.config import EvaluationConfig
from latentrouter.data.splits import create_splits
from latentrouter.io import ensure_dir, read_json, stable_hash, write_json
from latentrouter.schemas import NormalizedArtifacts, RouterDatasetBundle

DEFAULT_BETA = 0.1
TSV_IMAGE_PREFIX = "tsvref::"

DATASET_METADATA: dict[str, dict[str, str]] = {
    "MMMU_DEV_VAL": {"task": "vqa_mc", "group_id": "general"},
    "MathVista_MINI": {"task": "math", "group_id": "stem"},
    "MathVision_MINI": {"task": "math", "group_id": "stem"},
    "MathVerse_MINI": {"task": "math", "group_id": "stem"},
    "MMBench_DEV_EN_V11": {"task": "vqa_mc", "group_id": "general"},
    "RealWorldQA": {"task": "vqa_oe", "group_id": "general"},
    "MMStar": {"task": "vqa_mc", "group_id": "general"},
    "HallusionBench": {"task": "vqa_mc", "group_id": "general"},
    "TextVQA_VAL": {"task": "ocr_qa", "group_id": "charts_ocr"},
    "ChartQA_TEST": {"task": "ocr_qa", "group_id": "charts_ocr"},
    "DocVQA_VAL": {"task": "ocr_qa", "group_id": "charts_ocr"},
    "InfoVQA_VAL": {"task": "ocr_qa", "group_id": "general"},
    "AI2D_TEST": {"task": "vqa_mc", "group_id": "stem"},
    "OCRBench": {"task": "ocr_qa", "group_id": "charts_ocr"},
}

MODEL_PRICING: dict[str, tuple[float, float]] = {
    "InternVL2_5-78B": (1.0, 1.5),
    "MiMo-VL-7B-RL": (0.2, 0.3),
    "Phi-3.5-Vision": (0.1, 0.1),
    "Qwen2.5-VL-32B-Instruct": (0.4, 0.6),
    "Qwen2.5-VL-72B-Instruct": (0.8, 1.2),
    "SmolVLM2": (0.06, 0.06),
    "llava_next_vicuna_7b": (0.2, 0.2),
    "Gemma3-27B": (0.35, 0.5),
    "Janus-Pro-1B": (0.05, 0.05),
    "Janus-Pro-7B": (0.18, 0.25),
    "Kimi-VL-A3B-Thinking-2506": (0.2, 0.25),
    "Pixtral-12B": (0.25, 0.35),
    "Qianfan-VL-8B": (0.18, 0.25),
    "deepseek_vl2": (0.35, 0.5),
    "deepseek_vl2_tiny": (0.05, 0.05),
    "GPT4o": (2.5, 10.0),
    "GeminiFlash2-5": (0.3, 2.4),
}


def rank_score(
    avg_acc: float,
    avg_cost: float,
    cmin: float,
    cmax: float,
    beta: float = DEFAULT_BETA,
    eps: float = 1e-12,
) -> float:
    cost_clipped = float(np.clip(avg_cost, cmin, cmax))
    log_cmax = np.log2(cmax)
    log_cmin = np.log2(cmin)
    log_cost = np.log2(cost_clipped)
    cost_term = (log_cmax - log_cost) / (log_cmax - log_cmin + eps)
    cost_term = float(np.clip(cost_term, 0.0, 1.0))
    accuracy = float(np.clip(avg_acc, 0.0, 1.0))
    score = (1 + beta) * accuracy * cost_term / (beta * accuracy + cost_term + eps)
    return float(np.clip(score, 0.0, 1.0))


def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix == ".parquet":
        return pd.read_parquet(path)
    if path.suffix == ".json":
        return pd.read_json(path)
    if path.suffix == ".jsonl":
        return pd.read_json(path, lines=True)
    return pd.read_csv(path)


def _find_first(root: Path, candidates: list[str]) -> Path | None:
    for relative in candidates:
        candidate = root / relative
        if candidate.exists():
            return candidate
    return None


def _load_matrix(path: Path) -> np.ndarray:
    if path.suffix == ".npz":
        archive = np.load(path, allow_pickle=True)
        for key in ("Y", "C", "quality", "cost", "arr_0"):
            if key in archive:
                return np.asarray(archive[key], dtype=float)
        first_key = next(iter(archive.files))
        return np.asarray(archive[first_key], dtype=float)
    return np.asarray(np.load(path, allow_pickle=True), dtype=float)


def _resolve_source_root(source: str, processed_dir: Path) -> Path:
    source_path = Path(source)
    if not source_path.exists():
        raise FileNotFoundError(f"Source path does not exist: {source}")
    if source_path.is_dir():
        return source_path
    if source_path.suffixes[-2:] == [".tar", ".gz"] or source_path.suffix == ".tgz":
        extracted_name = source_path.name.removesuffix(".tar.gz")
        sibling_extracted = source_path.parent / extracted_name
        if sibling_extracted.exists():
            return sibling_extracted
        extract_root = ensure_dir(processed_dir / "_vl_routerbench_source" / stable_hash(str(source_path.resolve()))[:16])
        if not any(extract_root.iterdir()):
            with tarfile.open(source_path, "r:*") as archive:
                archive.extractall(extract_root)
        nested = extract_root / "vlm_router_data"
        return nested if nested.exists() else extract_root
    raise ValueError("VL-RouterBench source must be a directory or a .tar.gz archive.")


def _normalize_image_paths(value: Any) -> list[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    if isinstance(value, list):
        return [str(item) for item in value if str(item)]
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return []
        if text.startswith("["):
            try:
                decoded = json.loads(text)
                if isinstance(decoded, list):
                    return [str(item) for item in decoded if str(item)]
            except json.JSONDecodeError:
                pass
        if "|" in text:
            return [part for part in text.split("|") if part]
        return [text]
    return [str(value)]


def _extract_models_from_pickle(path: Path, expected_count: int) -> list[dict[str, Any]]:
    with path.open("rb") as handle:
        payload = pickle.load(handle)
    if isinstance(payload, dict):
        items = sorted(payload.items(), key=lambda item: int(item[1]) if isinstance(item[1], int) else item[0])
        return [{"model_id": str(model_id), "index": int(index) if isinstance(index, int) else idx} for idx, (model_id, index) in enumerate(items)]
    if isinstance(payload, (list, tuple)):
        return [{"model_id": str(model_id), "index": idx} for idx, model_id in enumerate(payload)]
    raise ValueError(f"Unsupported model index payload: {type(payload)!r}")


def _load_model_registry(registry_dir: Path, expected_count: int) -> list[dict[str, Any]]:
    for name in ("models.parquet", "models.csv", "models.json"):
        path = registry_dir / name
        if not path.exists():
            continue
        frame = _read_table(path)
        if "model_id" not in frame.columns:
            continue
        records = frame.to_dict(orient="records")
        return [
            {
                "model_id": str(record["model_id"]),
                **{key: value for key, value in record.items() if key != "model_id"},
            }
            for record in records
        ]

    model_index_path = _find_first(registry_dir, ["model_index.pkl", "model_ids.pkl"])
    if model_index_path is not None:
        return _extract_models_from_pickle(model_index_path, expected_count)

    return [{"model_id": f"model_{idx:03d}", "index": idx} for idx in range(expected_count)]


def _canonical_value(record: dict[str, Any], *candidates: str, default: str = "") -> str:
    for key in candidates:
        value = record.get(key)
        if value is not None and not (isinstance(value, float) and pd.isna(value)):
            text = str(value)
            if text:
                return text
    return default


def _load_splits(
    root: Path,
    samples: pd.DataFrame,
    split_manifest: str | Path | None,
    seed: int,
    train_fraction: float,
    validation_fraction: float,
    test_fraction: float,
) -> tuple[pd.DataFrame, str]:
    if split_manifest:
        split_frame = create_splits(
            samples,
            split_manifest=split_manifest,
            seed=seed,
            train_fraction=train_fraction,
            validation_fraction=validation_fraction,
            test_fraction=test_fraction,
            group_columns=["group_id", "dataset_name"],
        )
        return split_frame, "manifest"
    split_dir = root / "SPLITS"
    if split_dir.exists():
        rows: list[dict[str, str]] = []
        for split_name, aliases in {
            "train": ["train", "training"],
            "val": ["dev", "val", "valid", "validation"],
            "test": ["test"],
        }.items():
            found = None
            for alias in aliases:
                for suffix in (".jsonl", ".json", ".csv", ".parquet"):
                    candidate = split_dir / f"{alias}{suffix}"
                    if candidate.exists():
                        found = candidate
                        break
                if found is not None:
                    break
            if found is None:
                continue
            frame = _read_table(found)
            if "sample_id" not in frame.columns:
                if "id" in frame.columns:
                    frame = frame.rename(columns={"id": "sample_id"})
                else:
                    raise ValueError(f"Unsupported split file without sample_id column: {found}")
            rows.extend({"sample_id": str(sample_id), "split": split_name} for sample_id in frame["sample_id"].astype(str))
        if rows:
            split_frame = pd.DataFrame(rows)
            return samples[["sample_id"]].merge(split_frame, on="sample_id", how="left"), "official"

    if "split" in samples.columns:
        split_frame = samples[["sample_id", "split"]].copy()
        split_frame["split"] = split_frame["split"].replace({"dev": "val", "valid": "val", "validation": "val"})
        return split_frame, "metadata"

    split_frame = create_splits(
        samples,
        split_manifest=None,
        seed=seed,
        train_fraction=train_fraction,
        validation_fraction=validation_fraction,
        test_fraction=test_fraction,
        group_columns=["group_id", "dataset_name"],
    )
    return split_frame, "generated"


def _load_cost_bounds(path: Path | None, costs: np.ndarray) -> dict[str, float]:
    if path is not None and path.exists():
        payload = read_json(path)
        if {"cmin", "cmax"}.issubset(payload):
            return {"cmin": float(payload["cmin"]), "cmax": float(payload["cmax"]), "beta": float(payload.get("beta", DEFAULT_BETA))}
    finite_costs = costs[np.isfinite(costs)]
    if finite_costs.size == 0:
        return {"cmin": 1e-6, "cmax": 1.0, "beta": DEFAULT_BETA}
    return {
        "cmin": float(np.nanmin(finite_costs)),
        "cmax": float(np.nanmax(finite_costs)),
        "beta": DEFAULT_BETA,
    }


def _iter_workbook_rows(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(rows)]
    output: list[dict[str, Any]] = []
    for row in rows:
        output.append(dict(zip(header, row)))
    workbook.close()
    return output


def _coerce_quality(value: Any) -> float:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return np.nan
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, np.integer, float, np.floating)):
        return float(value)
    text = str(value).strip().lower()
    if text in {"true", "yes", "correct"}:
        return 1.0
    if text in {"false", "no", "incorrect"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return np.nan


def _approx_token_count(text: str) -> int:
    normalized = str(text or "").strip()
    if not normalized:
        return 0
    return len(re.findall(r"\w+|[^\w\s]", normalized))


def _build_prompt_text(record: dict[str, Any]) -> str:
    parts: list[str] = []
    for key in ("question", "instruction", "prompt"):
        value = record.get(key)
        if value is not None and str(value).strip():
            parts.append(str(value).strip())
            break
    hint = record.get("hint")
    if hint is not None and str(hint).strip():
        parts.append(f"Hint: {str(hint).strip()}")
    for option_key in ("A", "B", "C", "D", "E", "F"):
        value = record.get(option_key)
        if value is not None and str(value).strip():
            parts.append(f"{option_key}: {str(value).strip()}")
    return "\n".join(parts).strip()


def _parse_dataset_name_from_workbook(model_id: str, workbook_path: Path) -> str:
    stem = workbook_path.stem
    prefix = f"{model_id}_"
    remainder = stem[len(prefix):] if stem.startswith(prefix) else stem
    for dataset_name in sorted(DATASET_METADATA, key=len, reverse=True):
        if remainder.startswith(dataset_name):
            return dataset_name
    raise ValueError(f"Could not infer dataset name from workbook: {workbook_path.name}")


def _build_tsv_image_ref(source_root: Path, dataset_name: str, index_value: str) -> str:
    return f"{TSV_IMAGE_PREFIX}{(source_root / 'TSV_images' / f'{dataset_name}.tsv').resolve()}::index::{index_value}"


def _build_raw_model_registry(model_ids: list[str]) -> list[dict[str, Any]]:
    registry = []
    for index, model_id in enumerate(model_ids):
        input_price, output_price = MODEL_PRICING.get(model_id, (0.1, 0.1))
        registry.append(
            {
                "model_id": model_id,
                "index": index,
                "input_price_per_1m": input_price,
                "output_price_per_1m": output_price,
            }
        )
    return registry


def _prepare_from_precomputed_layout(
    source: str,
    source_root: Path,
    processed_dir: Path,
    split_manifest: str | Path | None,
    train_fraction: float,
    validation_fraction: float,
    test_fraction: float,
    seed: int,
) -> NormalizedArtifacts:
    matrices_dir = source_root / "data" / "matrices"
    registry_dir = source_root / "data" / "registry"

    quality_path = _find_first(matrices_dir, ["Y.npz", "Y.npy", "quality.npy", "quality.npz"])
    cost_path = _find_first(matrices_dir, ["C.npy", "C.npz", "cost.npy", "cost.npz"])
    meta_path = _find_first(registry_dir, ["meta.parquet", "meta.csv", "meta.json", "meta.jsonl"])
    if quality_path is None or cost_path is None or meta_path is None:
        raise ValueError(
            "Unsupported VL-RouterBench layout. Expected prepared matrices under data/matrices "
            "and sample metadata under data/registry/meta.*."
        )

    quality_matrix = _load_matrix(quality_path)
    cost_matrix = _load_matrix(cost_path)
    if cost_matrix.ndim == 1:
        cost_matrix = np.broadcast_to(cost_matrix.reshape(1, -1), quality_matrix.shape).copy()
    if quality_matrix.shape != cost_matrix.shape:
        raise ValueError(
            f"Quality and cost matrices must match. Got {quality_matrix.shape} vs {cost_matrix.shape}."
        )

    token_path = _find_first(matrices_dir, ["T.npy", "T.npz", "token_count.npy", "token_count.npz", "tokens.npy"])
    token_matrix = _load_matrix(token_path) if token_path is not None else None
    if token_matrix is not None and token_matrix.ndim == 1:
        token_matrix = np.broadcast_to(token_matrix.reshape(1, -1), quality_matrix.shape).copy()

    meta = _read_table(meta_path).reset_index(drop=True)
    if len(meta) != quality_matrix.shape[0]:
        raise ValueError(
            f"Metadata rows must match matrix rows. Got {len(meta)} rows for matrix height {quality_matrix.shape[0]}."
        )

    model_registry = _load_model_registry(registry_dir, expected_count=quality_matrix.shape[1])
    if len(model_registry) != quality_matrix.shape[1]:
        raise ValueError(
            f"Model registry length must match matrix width. Got {len(model_registry)} models for {quality_matrix.shape[1]} columns."
        )
    model_ids = [str(item["model_id"]) for item in model_registry]

    sample_rows: list[dict[str, Any]] = []
    outcome_rows: list[dict[str, Any]] = []
    for row_idx, row in meta.iterrows():
        record = row.to_dict()
        sample_id = _canonical_value(record, "sample_id", "id", default=stable_hash(f"vl::{row_idx}"))
        dataset_name = _canonical_value(record, "dataset_name", "dataset", "benchmark", "task", default="unknown")
        group_id = _canonical_value(record, "group_id", "task_group", "family", "category", default=dataset_name)
        mode_id = _canonical_value(record, "mode_id", "mode", "subset", "task_type", default="default")
        question = _canonical_value(record, "question", "query", "prompt", "instruction", "text")
        answer = _canonical_value(record, "answer", "ground_truth", "label", "gt")
        image_paths = _normalize_image_paths(
            record.get("image_paths") or record.get("images") or record.get("image_path") or record.get("image")
        )
        metadata = {
            key: value
            for key, value in record.items()
            if key not in {"sample_id", "id", "dataset_name", "dataset", "benchmark", "task", "group_id", "task_group", "family", "category", "mode_id", "mode", "subset", "task_type", "question", "query", "prompt", "instruction", "text", "answer", "ground_truth", "label", "gt", "image_paths", "images", "image_path", "image", "split"}
            and value is not None
            and not (isinstance(value, float) and pd.isna(value))
        }
        sample_rows.append(
            {
                "sample_id": sample_id,
                "benchmark_id": "vl_routerbench",
                "dataset_idx": f"{dataset_name}::{mode_id}",
                "group_id": group_id,
                "dataset_name": dataset_name,
                "mode_id": mode_id,
                "question": question,
                "answer": answer,
                "image_paths": image_paths,
                "image_count": len(image_paths),
                "split": record.get("split"),
                "metadata_json": json.dumps(metadata, sort_keys=True),
            }
        )
        for model_idx, model_id in enumerate(model_ids):
            quality = float(quality_matrix[row_idx, model_idx]) if np.isfinite(quality_matrix[row_idx, model_idx]) else np.nan
            cost = float(cost_matrix[row_idx, model_idx]) if np.isfinite(cost_matrix[row_idx, model_idx]) else np.nan
            token_count = np.nan
            if token_matrix is not None:
                token_value = token_matrix[row_idx, model_idx]
                token_count = float(token_value) if np.isfinite(token_value) else np.nan
            if np.isnan(quality) and np.isnan(cost):
                continue
            outcome_rows.append(
                {
                    "sample_id": sample_id,
                    "model_id": model_id,
                    "quality": quality,
                    "prediction": "",
                    "is_correct": quality,
                    "token_count": token_count,
                    "prompt_tokens": np.nan,
                    "completion_tokens": np.nan,
                    "cost": cost,
                    "raw_output": None,
                    "metadata_json": json.dumps({}, sort_keys=True),
                }
            )

    samples = pd.DataFrame(sample_rows).sort_values("sample_id").reset_index(drop=True)
    outcomes = pd.DataFrame(outcome_rows).sort_values(["sample_id", "model_id"]).reset_index(drop=True)
    splits, split_policy = _load_splits(
        source_root,
        samples,
        split_manifest=split_manifest,
        seed=seed,
        train_fraction=train_fraction,
        validation_fraction=validation_fraction,
        test_fraction=test_fraction,
    )
    cost_bounds = _load_cost_bounds(_find_first(matrices_dir, ["cost_bounds.json"]), cost_matrix)

    return _write_prepared_outputs(
        source=source,
        source_root=source_root,
        processed_dir=processed_dir,
        samples=samples,
        outcomes=outcomes,
        model_registry=model_registry,
        splits=splits,
        split_policy=split_policy,
        cost_bounds=cost_bounds,
        source_layout="precomputed",
        cost_mode="provided",
    )


def _prepare_from_raw_layout(
    source: str,
    source_root: Path,
    processed_dir: Path,
    split_manifest: str | Path | None,
    train_fraction: float,
    validation_fraction: float,
    test_fraction: float,
    seed: int,
) -> NormalizedArtifacts:
    evaluation_root = source_root / "VLMEvalKit_evaluation"
    if not evaluation_root.exists():
        raise ValueError("Raw VL-RouterBench layout requires VLMEvalKit_evaluation/")

    model_dirs = sorted([path for path in evaluation_root.iterdir() if path.is_dir()], key=lambda path: path.name.lower())
    model_ids = [path.name for path in model_dirs]
    model_registry = _build_raw_model_registry(model_ids)

    sample_rows: dict[str, dict[str, Any]] = {}
    outcome_rows: list[dict[str, Any]] = []
    for model_dir in model_dirs:
        model_id = model_dir.name
        input_price, output_price = MODEL_PRICING.get(model_id, (0.1, 0.1))
        for workbook_path in sorted(model_dir.glob("*.xlsx")):
            dataset_name = _parse_dataset_name_from_workbook(model_id, workbook_path)
            dataset_info = DATASET_METADATA.get(dataset_name, {"task": "default", "group_id": "unknown"})
            for row_idx, record in enumerate(_iter_workbook_rows(workbook_path)):
                raw_index = _canonical_value(record, "index", "id", default=str(row_idx))
                sample_id = f"{dataset_name}::{raw_index}"
                prompt = _build_prompt_text(record)
                answer = _canonical_value(record, "answer")
                image_paths = []
                tsv_path = source_root / "TSV_images" / f"{dataset_name}.tsv"
                if tsv_path.exists():
                    image_paths = [_build_tsv_image_ref(source_root, dataset_name, raw_index)]
                elif record.get("image_path") is not None:
                    image_paths = [str(record["image_path"])]

                sample_rows.setdefault(
                    sample_id,
                    {
                        "sample_id": sample_id,
                        "benchmark_id": "vl_routerbench",
                        "dataset_idx": f"{dataset_name}::{dataset_info['task']}",
                        "group_id": dataset_info["group_id"],
                        "dataset_name": dataset_name,
                        "mode_id": dataset_info["task"],
                        "question": prompt,
                        "answer": answer,
                        "image_paths": image_paths,
                        "image_count": len(image_paths),
                        "metadata_json": json.dumps(
                            {
                                key: value
                                for key, value in record.items()
                                if key not in {"prediction", "is_correct", "score", "hit", "log", "eval_gt", "eval_pred", "eval_match", "eval_score"}
                                and value is not None
                            },
                            sort_keys=True,
                        ),
                    },
                )

                quality = _coerce_quality(record.get("is_correct"))
                if np.isnan(quality):
                    quality = _coerce_quality(record.get("score"))
                prediction = _canonical_value(record, "prediction", "eval_pred")
                prompt_tokens = float(_approx_token_count(prompt))
                completion_tokens = float(_approx_token_count(prediction))
                token_count = prompt_tokens + completion_tokens
                cost = (prompt_tokens * input_price + completion_tokens * output_price) / 1_000_000.0
                outcome_rows.append(
                    {
                        "sample_id": sample_id,
                        "model_id": model_id,
                        "quality": quality,
                        "prediction": prediction,
                        "is_correct": quality,
                        "token_count": token_count,
                        "prompt_tokens": prompt_tokens,
                        "completion_tokens": completion_tokens,
                        "cost": cost,
                        "raw_output": prediction,
                        "metadata_json": json.dumps({}, sort_keys=True),
                    }
                )

    samples = pd.DataFrame(sample_rows.values()).sort_values("sample_id").reset_index(drop=True)
    outcomes = pd.DataFrame(outcome_rows).sort_values(["sample_id", "model_id"]).reset_index(drop=True)
    cost_bounds = _load_cost_bounds(None, outcomes["cost"].to_numpy(dtype=float))
    splits, split_policy = _load_splits(
        source_root,
        samples,
        split_manifest=split_manifest,
        seed=seed,
        train_fraction=train_fraction,
        validation_fraction=validation_fraction,
        test_fraction=test_fraction,
    )
    return _write_prepared_outputs(
        source=source,
        source_root=source_root,
        processed_dir=processed_dir,
        samples=samples,
        outcomes=outcomes,
        model_registry=model_registry,
        splits=splits,
        split_policy=split_policy,
        cost_bounds=cost_bounds,
        source_layout="raw_vlmevalkit_archive",
        cost_mode="approx_text_tokens",
    )


def _write_prepared_outputs(
    source: str,
    source_root: Path,
    processed_dir: Path,
    samples: pd.DataFrame,
    outcomes: pd.DataFrame,
    model_registry: list[dict[str, Any]],
    splits: pd.DataFrame,
    split_policy: str,
    cost_bounds: dict[str, float],
    source_layout: str,
    cost_mode: str,
) -> NormalizedArtifacts:
    samples_path = processed_dir / "samples.parquet"
    outcomes_path = processed_dir / "outcomes.parquet"
    models_path = processed_dir / "models.json"
    manifest_path = processed_dir / "manifest.json"
    splits_path = processed_dir / "splits.parquet"
    samples.to_parquet(samples_path, index=False)
    outcomes.to_parquet(outcomes_path, index=False)
    splits.to_parquet(splits_path, index=False)
    write_json(models_path, model_registry)
    write_json(
        manifest_path,
        {
            "adapter_version": 3,
            "benchmark_id": "vl_routerbench",
            "protocol_id": "vl_routerbench_rank_score",
            "source": source,
            "source_layout_root": str(source_root),
            "source_layout": source_layout,
            "split_policy": split_policy,
            "sample_count": int(len(samples)),
            "outcome_count": int(len(outcomes)),
            "model_count": int(len(model_registry)),
            "cost_bounds": cost_bounds,
            "cost_mode": cost_mode,
            "image_reference_mode": "tsvref",
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
    )
    return NormalizedArtifacts(
        processed_dir=processed_dir,
        samples_path=samples_path,
        outcomes_path=outcomes_path,
        models_path=models_path,
        manifest_path=manifest_path,
        splits_path=splits_path,
    )


class VLRouterBenchAdapter(BenchmarkAdapter):
    benchmark_id = "vl_routerbench"
    protocol_id = "vl_routerbench_rank_score"

    def prepare(
        self,
        source: str,
        processed_dir: str | Path,
        source_split: str = "train",
        split_manifest: str | Path | None = None,
        train_fraction: float = 0.7,
        validation_fraction: float = 0.1,
        test_fraction: float = 0.2,
        seed: int = 20260308,
        adapter_options: dict[str, Any] | None = None,
    ) -> NormalizedArtifacts:
        del source_split, adapter_options
        processed_dir = ensure_dir(processed_dir)
        source_root = _resolve_source_root(source, processed_dir)
        matrices_dir = source_root / "data" / "matrices"
        registry_dir = source_root / "data" / "registry"
        if matrices_dir.exists() and registry_dir.exists():
            return _prepare_from_precomputed_layout(
                source=source,
                source_root=source_root,
                processed_dir=processed_dir,
                split_manifest=split_manifest,
                train_fraction=train_fraction,
                validation_fraction=validation_fraction,
                test_fraction=test_fraction,
                seed=seed,
            )
        if (source_root / "VLMEvalKit_evaluation").exists():
            return _prepare_from_raw_layout(
                source=source,
                source_root=source_root,
                processed_dir=processed_dir,
                split_manifest=split_manifest,
                train_fraction=train_fraction,
                validation_fraction=validation_fraction,
                test_fraction=test_fraction,
                seed=seed,
            )
        raise ValueError(
            "Unsupported VL-RouterBench layout. Expected either prepared matrices under data/matrices "
            "or the raw archive layout with VLMEvalKit_evaluation/ and TSV_images/."
        )


def _vl_frontier_metrics(
    frontier: pd.DataFrame,
    cost_bounds: dict[str, float],
    sample_count: int,
    prediction_wall_time_seconds: float | None,
) -> pd.DataFrame:
    beta = float(cost_bounds.get("beta", DEFAULT_BETA))
    scored = frontier.copy()
    scored["avg_accuracy"] = scored["mean_quality"].astype(float)
    scored["avg_cost"] = scored["mean_cost"].astype(float)
    scored["rank_score"] = [
        rank_score(acc, cost, cost_bounds["cmin"], cost_bounds["cmax"], beta=beta)
        for acc, cost in zip(scored["avg_accuracy"], scored["avg_cost"])
    ]
    if prediction_wall_time_seconds and prediction_wall_time_seconds > 0 and "mean_token_count" in scored.columns:
        scored["throughput_k_tokens_per_s"] = (
            scored["mean_token_count"] * float(sample_count) / (prediction_wall_time_seconds * 1000.0)
        )
    return scored


def _vl_slice_metrics(
    routes: pd.DataFrame,
    aggregate_by: list[str],
    cost_bounds: dict[str, float],
    sample_count_lookup: dict[tuple[str, str], int],
    prediction_wall_time_seconds: float | None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for field in aggregate_by:
        if field not in routes.columns:
            continue
        for value, frame in routes.groupby(field, dropna=False):
            aggregations: dict[str, tuple[str, str]] = {
                "mean_cost": ("cost", "mean"),
                "mean_quality": ("correctness", "mean"),
            }
            if "token_count" in frame.columns:
                aggregations["mean_token_count"] = ("token_count", "mean")
            frontier = frame.groupby("lambda_value", as_index=False).agg(**aggregations).sort_values("mean_cost").reset_index(drop=True)
            frontier = _vl_frontier_metrics(
                frontier,
                cost_bounds,
                sample_count_lookup.get((field, str(value)), int(frame["sample_id"].nunique())),
                prediction_wall_time_seconds,
            )
            best_row = frontier.sort_values(["rank_score", "avg_accuracy"], ascending=[False, False]).iloc[0]
            row: dict[str, Any] = {
                "slice_field": field,
                "slice_value": value,
                "primary_metric_value": float(best_row["rank_score"]),
                "avg_accuracy": float(best_row["avg_accuracy"]),
                "avg_cost": float(best_row["avg_cost"]),
                "rank_score": float(best_row["rank_score"]),
                "sample_count": int(frame["sample_id"].nunique()),
            }
            if "throughput_k_tokens_per_s" in best_row:
                row["throughput_k_tokens_per_s"] = float(best_row["throughput_k_tokens_per_s"])
            rows.append(row)
    return pd.DataFrame(rows)


class VLRouterBenchProtocol(BenchmarkProtocol):
    benchmark_id = "vl_routerbench"
    protocol_id = "vl_routerbench_rank_score"

    def evaluate(
        self,
        router_name: str,
        predicted_utilities: np.ndarray,
        bundle: RouterDatasetBundle,
        config: EvaluationConfig,
        run_dir: str | Path,
        prediction_wall_time_seconds: float | None = None,
    ) -> dict[str, Any]:
        lambdas = lambda_grid(config)
        routes, frontier = materialize_routes(predicted_utilities, bundle, lambdas)
        single_model = single_model_frontier(bundle)
        oracle = oracle_frontier(bundle, lambdas)
        cost_bounds = {
            "cmin": float(bundle.evaluation_metadata.get("cost_bounds", {}).get("cmin", float(single_model["cost"].min()))),
            "cmax": float(bundle.evaluation_metadata.get("cost_bounds", {}).get("cmax", float(single_model["cost"].max()))),
            "beta": float(bundle.evaluation_metadata.get("cost_bounds", {}).get("beta", DEFAULT_BETA)),
        }
        frontier = _vl_frontier_metrics(
            frontier,
            cost_bounds,
            len(bundle.sample_ids),
            prediction_wall_time_seconds if config.measure_throughput else None,
        )
        oracle = _vl_frontier_metrics(
            oracle,
            cost_bounds,
            len(bundle.sample_ids),
            prediction_wall_time_seconds if config.measure_throughput else None,
        )
        single_model = single_model.copy()
        single_model["avg_accuracy"] = single_model["quality"]
        single_model["avg_cost"] = single_model["cost"]
        single_model["rank_score"] = [
            rank_score(acc, cost, cost_bounds["cmin"], cost_bounds["cmax"], beta=cost_bounds["beta"])
            for acc, cost in zip(single_model["avg_accuracy"], single_model["avg_cost"])
        ]
        slice_metrics = _vl_slice_metrics(
            routes.rename(columns={"selected_model": "model_id"}),
            config.aggregate_by,
            cost_bounds,
            {
                (field, str(value)): int(frame["sample_id"].nunique())
                for field in config.aggregate_by
                if field in routes.columns
                for value, frame in routes.groupby(field, dropna=False)
            },
            prediction_wall_time_seconds if config.measure_throughput else None,
        )
        best_row = frontier.sort_values(["rank_score", "avg_accuracy"], ascending=[False, False]).iloc[0]
        best_single = single_model.sort_values(["rank_score", "avg_accuracy"], ascending=[False, False]).iloc[0]

        metrics = {
            "benchmark_id": self.benchmark_id,
            "protocol_id": self.protocol_id,
            "router_name": router_name,
            "primary_metric": "rank_score",
            "primary_metric_value": float(best_row["rank_score"]),
            "summary_metric_order": ["avg_accuracy", "avg_cost", "rank_score", "best_lambda_value", "best_single_model", "best_single_rank_score"],
            "slice_metric_order": ["avg_accuracy", "avg_cost", "rank_score"],
            "avg_accuracy": float(best_row["avg_accuracy"]),
            "avg_cost": float(best_row["avg_cost"]),
            "rank_score": float(best_row["rank_score"]),
            "best_lambda_value": float(best_row["lambda_value"]),
            "best_single_model": str(best_single["model_id"]),
            "best_single_rank_score": float(best_single["rank_score"]),
            "cost_bounds_cmin": float(cost_bounds["cmin"]),
            "cost_bounds_cmax": float(cost_bounds["cmax"]),
            "rank_score_beta": float(cost_bounds["beta"]),
            "lambda_count": int(len(lambdas)),
            "sample_count": int(len(bundle.sample_ids)),
            "model_count": int(len(bundle.model_ids)),
            "plot_x_label": "Average cost",
            "plot_y_label": "Average accuracy",
            "plot_title": "Accuracy-cost frontier",
        }
        if config.measure_throughput and "throughput_k_tokens_per_s" in best_row:
            metrics["throughput_k_tokens_per_s"] = float(best_row["throughput_k_tokens_per_s"])
            metrics["summary_metric_order"] = metrics["summary_metric_order"] + ["throughput_k_tokens_per_s"]

        return write_evaluation_outputs(run_dir, metrics, frontier, routes, single_model, oracle, slice_metrics)
